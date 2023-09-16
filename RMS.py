from tkinter import *
import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
from tkinter import messagebox
import random
import time



root = Tk()
root.geometry("1600x700")
#root.resizable(0,0)
root.title("Restaurant Management System")
root.config(background="#080808")

Tops = Frame(root,bg="yellow",width = 1600,height=50)
Tops.pack(side=TOP)

f1 = Frame(root,width = 900,height=700,bg="#080808")
f1.pack(side=LEFT)

f2 = Frame(root ,width = 400,height=700,bg="#080808")
f2.pack(side=RIGHT)
#TIME
localtime=time.asctime(time.localtime(time.time()))


lblinfo = Label(Tops, font=( 'aria' ,30, 'bold' ),text="Restaurant Management System",fg="yellow",bd=10,background="#080808")
lblinfo.grid(row=0,column=0)
lblinfo = Label(Tops, font=( 'aria' ,20, ),text=localtime,fg="Yellow",bg="#080808")
lblinfo.grid(row=1,column=0)

#Calculator
text_Input=StringVar()
operator =""

txtdisplay = Entry(f2,font=('ariel' ,20,'bold'), textvariable=text_Input , bd=5 ,insertwidth=5 ,bg="white",justify='right')
txtdisplay.grid(columnspan=4)

def  btnclick(numbers):
    global operator
    operator=operator + str(numbers)
    text_Input.set(operator)

def clrdisplay():
    global operator
    operator=""
    text_Input.set("")

def eqals():
    global operator
    sumup=str(eval(operator))

    text_Input.set(sumup)
    operator = ""

def totalbill():

    cof =float(Fries.get())
    colfries= float(Largefries.get())
    cob= float(Burger.get())
    cofi= float(Filet.get())
    cochee= float(Cheese_burger.get())
    codr= float(Drinks.get())

    costoffries = cof*25
    costoflargefries = colfries*40
    costofburger = cob*35
    costoffilet = cofi*50
    costofcheeseburger = cochee*50
    costofdrinks = codr*35

    costofmeal = "Rs.",str('%.2f'% (costoffries +  costoflargefries + costofburger + costoffilet + costofcheeseburger + costofdrinks))
    PayTax=((costoffries +  costoflargefries + costofburger + costoffilet +  costofcheeseburger + costofdrinks)*0.33)
    Totalcost=(costoffries +  costoflargefries + costofburger + costoffilet  + costofcheeseburger + costofdrinks)
    Ser_Charge=((costoffries +  costoflargefries + costofburger + costoffilet + costofcheeseburger + costofdrinks)/99)
    Service="Rs.",str('%.2f'% Ser_Charge)
    OverAllCost="Rs.",str('%.2f'%(PayTax + Totalcost + Ser_Charge))
    PaidTax="Rs.",str('%.2f'% PayTax)

    Service_Charge.set(Service)
    cost.set(costofmeal)
    Tax.set(PaidTax)
    Subtotal.set(costofmeal)
    Total.set(OverAllCost)
    
file=pathlib.Path("Bill_Data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Date & Time"
    sheet["B1"]="Bill NO."
    sheet["C1"]="Fries Meal"
    sheet["D1"]="Lunch Meal"
    sheet["E1"]="Burger"
    sheet["F1"]="Pizza"
    sheet["G1"]="Cheese Burger"
    sheet["H1"]="Cold Drink"
    sheet["I1"]="Cost"
    sheet["J1"]="Service Charge"
    sheet["K1"]="Tax"
    sheet["L1"]="Subtotal"
    sheet["M1"]="Total"
    
    file.save("Bill_Data.xlsx")


def save_data():
    
    bill=txtreference.get() 
    fries=txtfries.get()
    lunch=txtLargefries.get()
    burger=txtburger.get()
    pizza=txtburger.get()
    cheez=txtCheese_burger.get()
    drink=txtDrinks.get()
    cost=txtcost.get()
    service=txtService_Charge.get()
    tax=txtTax.get()
    subtotal=txtSubtotal.get()
    total=txtTotal.get()
    
    if fries=="" or lunch=="" or burger=="" or pizza=="" or cheez=="" or drink=="":
        messagebox.showerror("Error","Please do not Leave any Field Empty...")
    else:
        ask=messagebox.askquestion("Generated Bill","Do You want to Save the Bill?")
        if ask=="yes":
            
            file=openpyxl.load_workbook("Bill_Data.xlsx")
            sheet=file.active
        
            sheet.cell(column=1,row=sheet.max_row+1,value=localtime)
            sheet.cell(column=2,row=sheet.max_row,value=bill)
            sheet.cell(column=3,row=sheet.max_row,value=fries)
            sheet.cell(column=4,row=sheet.max_row,value=lunch)
            sheet.cell(column=5,row=sheet.max_row,value=burger)
            sheet.cell(column=6,row=sheet.max_row,value=pizza)
            sheet.cell(column=7,row=sheet.max_row,value=cheez)
            sheet.cell(column=8,row=sheet.max_row,value=drink)
            sheet.cell(column=9,row=sheet.max_row,value=cost)
            sheet.cell(column=10,row=sheet.max_row,value=service)
            sheet.cell(column=11,row=sheet.max_row,value=tax)
            sheet.cell(column=12,row=sheet.max_row,value=subtotal)
            sheet.cell(column=13,row=sheet.max_row,value=total)
            
            
            file.save("Bill_Data.xlsx")
            messagebox.showinfo("Restaurant Management System","Your Data Have been Successfully Saved...")
                    
    
        
def reset():
    rand.set("")
    Fries.set("")
    Largefries.set("")
    Burger.set("")
    Filet.set("")
    Subtotal.set("")
    Total.set("")
    Service_Charge.set("")
    Drinks.set("")
    Tax.set("")
    cost.set("")
    Cheese_burger.set("")


btn7=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="7",bg="#080808", command=lambda: btnclick(7) )
btn7.grid(row=2,column=0)

btn8=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="8",bg="#080808", command=lambda: btnclick(8) )
btn8.grid(row=2,column=1)

btn9=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="9",bg="#080808", command=lambda: btnclick(9) )
btn9.grid(row=2,column=2)

Addition=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="+",bg="#080808", command=lambda: btnclick("+") )
Addition.grid(row=2,column=3)

btn4=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="4",bg="#080808", command=lambda: btnclick(4) )
btn4.grid(row=3,column=0)

btn5=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="5",bg="#080808", command=lambda: btnclick(5) )
btn5.grid(row=3,column=1)

btn6=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="6",bg="#080808", command=lambda: btnclick(6) )
btn6.grid(row=3,column=2)

Substraction=Button(f2,padx=19,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="-",bg="#080808", command=lambda: btnclick("-") )
Substraction.grid(row=3,column=3)

btn1=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="1",bg="#080808", command=lambda: btnclick(1) )
btn1.grid(row=4,column=0)

btn2=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="2",bg="#080808", command=lambda: btnclick(2) )
btn2.grid(row=4,column=1)

btn3=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="3",bg="#080808", command=lambda: btnclick(3) )
btn3.grid(row=4,column=2)

multiply=Button(f2,padx=19,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="*",bg="#080808", command=lambda: btnclick("*") )
multiply.grid(row=4,column=3)


btn0=Button(f2,padx=16,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="0",bg="#080808", command=lambda: btnclick(0) )
btn0.grid(row=5,column=0)

btnc=Button(f2,padx=16,pady=16,bd=4, fg="red", font=('ariel', 20 ,'bold'),text="c",bg="#080808", command=clrdisplay)
btnc.grid(row=5,column=1)

btnequal=Button(f2,padx=16,pady=16,bd=4,width = 16, fg="black", font=('ariel', 20 ,'bold'),text="=",bg="yellow",command=eqals)
btnequal.grid(columnspan=4)

Decimal=Button(f2,padx=19,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text=".",bg="#080808", command=lambda: btnclick(".") )
Decimal.grid(row=5,column=2)

Division=Button(f2,padx=20,pady=16,bd=4, fg="white", font=('ariel', 20 ,'bold'),text="/",bg="#080808", command=lambda: btnclick("/") )
Division.grid(row=5,column=3)
status = Label(f2,font=('aria', 15, 'bold'),width = 16, text="-By Satendra Baghel",bd=2,bg="#080808",fg="white")
status.grid(row=7,columnspan=3)


rand = StringVar()
Fries = StringVar()
Largefries = StringVar()
Burger = StringVar()
Filet = StringVar()
Subtotal = StringVar()
Total = StringVar()
Service_Charge = StringVar()
Drinks = StringVar()
Tax = StringVar()
cost = StringVar()
Cheese_burger = StringVar()

def generate_bill_no():
            generated_bill=''
            for r in range(6):
                generated_bill+= str(random.randint(0,9))  
            print(generated_bill)
            txtreference.config(state=NORMAL)
            txtreference.delete(0,END)
            txtreference.insert(END,generated_bill)
            
            
lblreference = Label(f1, font=( 'aria' ,16, 'bold' ),text="Bill No. :",fg="white",bd=10,bg="#080808")
lblreference.grid(row=0,column=0)
txtreference = Entry(f1,font=('ariel' ,16,'bold'), textvariable=rand , bd=6,insertwidth=4,bg="white" ,justify='right')
txtreference.grid(row=0,column=1)
generate_bill_no()

lblfries = Label(f1, font=( 'aria' ,16, 'bold' ),text="Fries Meal :",fg="white",bd=10,bg="#080808")
lblfries.grid(row=1,column=0)
txtfries = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Fries , bd=6,insertwidth=4,bg="white" ,justify='right')
txtfries.grid(row=1,column=1)

lblLargefries = Label(f1, font=( 'aria' ,16, 'bold' ),text="Lunch Meal :",fg="white",bd=10,bg="#080808")
lblLargefries.grid(row=2,column=0)
txtLargefries = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Largefries , bd=6,insertwidth=4,bg="white" ,justify='right')
txtLargefries.grid(row=2,column=1)


lblburger = Label(f1, font=( 'aria' ,16, 'bold' ),text="Burger :",fg="white",bd=10,bg="#080808")
lblburger.grid(row=3,column=0)
txtburger = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Burger , bd=6,insertwidth=4,bg="white" ,justify='right')
txtburger.grid(row=3,column=1)

lblFilet = Label(f1, font=( 'aria' ,16, 'bold' ),text="Pizza :",fg="white",bd=10,bg="#080808")
lblFilet.grid(row=4,column=0)
txtFilet = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Filet , bd=6,insertwidth=4,bg="white" ,justify='right')
txtFilet.grid(row=4,column=1)

lblCheese_burger = Label(f1, font=( 'aria' ,16, 'bold' ),text="Cheese burger :",fg="white",bd=10,bg="#080808")
lblCheese_burger.grid(row=5,column=0)
txtCheese_burger = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Cheese_burger , bd=6,insertwidth=4,bg="white" ,justify='right')
txtCheese_burger.grid(row=5,column=1)


lblDrinks = Label(f1, font=( 'aria' ,16, 'bold' ),text="Cold Drinks :",fg="white",bd=10,bg="#080808")
lblDrinks.grid(row=0,column=2)
txtDrinks = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Drinks , bd=6,insertwidth=4,bg="white" ,justify='right')
txtDrinks.grid(row=0,column=3)

lblcost = Label(f1, font=( 'aria' ,16, 'bold' ),text="cost :",fg="white",bd=10,bg="#080808")
lblcost.grid(row=1,column=2)
txtcost = Entry(f1,font=('ariel' ,16,'bold'), textvariable=cost , bd=6,insertwidth=4,bg="white" ,justify='right')
txtcost.grid(row=1,column=3)

lblService_Charge = Label(f1, font=( 'aria' ,16, 'bold' ),text="Service Charge :",fg="white",bd=10,bg="#080808")
lblService_Charge.grid(row=2,column=2)
txtService_Charge = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Service_Charge , bd=6,insertwidth=4,bg="white" ,justify='right')
txtService_Charge.grid(row=2,column=3)

lblTax = Label(f1, font=( 'aria' ,16, 'bold' ),text="Tax :",fg="white",bd=10,bg="#080808")
lblTax.grid(row=3,column=2)
txtTax = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Tax , bd=6,insertwidth=4,bg="white" ,justify='right')
txtTax.grid(row=3,column=3)

lblSubtotal = Label(f1, font=( 'aria' ,16, 'bold' ),text="Subtotal :",fg="white",bd=10,bg="#080808")
lblSubtotal.grid(row=4,column=2)
txtSubtotal = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Subtotal , bd=6,insertwidth=4,bg="white" ,justify='right')
txtSubtotal.grid(row=4,column=3)

lblTotal = Label(f1, font=( 'aria' ,16, 'bold' ),text="Total :",fg="white",bd=10,bg="#080808")
lblTotal.grid(row=5,column=2)
txtTotal = Entry(f1,font=('ariel' ,16,'bold'), textvariable=Total , bd=6,insertwidth=4,bg="white" ,justify='right')
txtTotal.grid(row=5,column=3)


lblTotal = Label(f1,text="---------------------",bg="#080808")
lblTotal.grid(row=6,columnspan=3)

btnTotal=Button(f1,padx=16,pady=8, bd=10 ,fg="white",font=('ariel' ,16,'bold'),width=10, text="TOTAL",bg="#080808",command=totalbill)
btnTotal.grid(row=7, column=1)

btnreset=Button(f1,padx=16,pady=8, bd=10 ,fg="white",font=('ariel' ,16,'bold'),width=10, text="RESET",bg="#080808",command=reset)
btnreset.grid(row=7, column=2)

btnsave=Button(f1,padx=16,pady=8, bd=10 ,fg="white",font=('ariel' ,16,'bold'),width=10, text="SAVE",bg="#080808",command=save_data)
btnsave.grid(row=7, column=3)

def price():
    roo = Tk()
    roo.geometry("600x220+0+0")
    roo.title("Price List")
    roo.config(bg="#080808")
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="ITEM", fg="white", bd=5,bg="#080808")
    lblinfo.grid(row=0, column=0)
    lblinfo = Label(roo, font=('aria', 15,'bold'), text="_____________", fg="white",bg="#080808")
    lblinfo.grid(row=0, column=2)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="PRICE", fg="white",bg="#080808")
    lblinfo.grid(row=0, column=3)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="Fries Meal", fg="white",bg="#080808")
    lblinfo.grid(row=1, column=0)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="25", fg="white",bg="#080808" )
    lblinfo.grid(row=1, column=3)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="Lunch Meal", fg="white",bg="#080808")
    lblinfo.grid(row=2, column=0)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="40", fg="white",bg="#080808")
    lblinfo.grid(row=2, column=3)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="Burger Meal", fg="white",bg="#080808")
    lblinfo.grid(row=3, column=0)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="35", fg="white",bg="#080808")
    lblinfo.grid(row=3, column=3)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="Pizza Meal", fg="white",bg="#080808")
    lblinfo.grid(row=4, column=0)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="50", fg="white",bg="#080808")
    lblinfo.grid(row=4, column=3)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="Cheese Burger", fg="white",bg="#080808")
    lblinfo.grid(row=5, column=0)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="30", fg="white",bg="#080808")
    lblinfo.grid(row=5, column=3)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="Cold Drinks", fg="white",bg="#080808")
    lblinfo.grid(row=6, column=0)
    lblinfo = Label(roo, font=('aria', 15, 'bold'), text="35", fg="white",bg="#080808")
    lblinfo.grid(row=6, column=3)

    roo.mainloop()

btnprice=Button(f1,padx=16,pady=8, bd=10 ,fg="white",font=('ariel' ,16,'bold'),width=10, text="PRICE",bg="#080808",command=price)
btnprice.grid(row=7, column=0)

root.mainloop()